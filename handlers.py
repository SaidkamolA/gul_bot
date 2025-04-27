from aiogram import Router, Bot, F
from aiogram.types import CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, Message
from aiogram.types.input_file import BufferedInputFile
from aiogram.filters import Command
import requests
from datetime import datetime, timedelta
from config import ADMIN_CHAT_ID, BACKEND_URL
from collections import defaultdict
import pandas as pd
import io
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList

router = Router()
sent_order_ids = set()  # To track sent orders and avoid duplicates
ORDERS_PER_PAGE = 5  # Number of orders to show per page

# List of admin IDs
ADMIN_IDS = {714948319, 6094832311, 575262312, ADMIN_CHAT_ID}  # Add your admin ID here

# Bot commands
COMMANDS = {
    'start': '🚀 Запустить бота и открыть админ-панель',
    'help': '❓ Показать список команд',
    'stats': '📊 Показать статистику',
    'orders': '📋 Показать все заказы',
    'pending': '⏳ Показать ожидающие заказы',
    'approved': '✅ Показать одобренные заказы',
    'rejected': '❌ Показать отклоненные заказы',
    'customers': '👥 Показать частых клиентов',
    'finance': '💰 Финансовая сводка',
    'products': '📦 Статистика по товарам',
    'download': '📥 Скачать полный отчет'
}

# Helper function to format timestamp
def format_timestamp(timestamp: str) -> str:
    try:
        dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
        return dt.strftime('%d.%m.%Y %H:%M:%S')
    except ValueError:
        return timestamp

# Helper function to get statistics
async def get_statistics():
    try:
        response = requests.get(BACKEND_URL)
        response.raise_for_status()
        orders = response.json()
        
        stats = {
            'total': len(orders),
            'approved': 0,
            'rejected': 0,
            'pending': 0,
            'total_quantity': 0,
            'products': defaultdict(int),
            'customers': defaultdict(int)  # Track customer orders
        }
        
        for order in orders:
            stats[order['status']] += 1
            stats['total_quantity'] += order['quantity']
            stats['products'][order['product']] += order['quantity']
            stats['customers'][order['phone']] += 1
            
        return stats
    except Exception as e:
        print(f"Error getting statistics: {e}")
        return None

# Helper function to apply styles to worksheet
def apply_styles(worksheet, title, data_start_row=2):
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Dark blue
    title_font = Font(bold=True, size=14, color="1F4E78")  # Dark blue
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue
    data_font = Font(size=11)
    border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # Apply title
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=worksheet.max_column)

    # Apply header styles
    for cell in worksheet[data_start_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Apply data styles
    for row in worksheet.iter_rows(min_row=data_start_row + 1, max_row=worksheet.max_row):
        for cell in row:
            cell.font = data_font
            cell.border = border
            # Align numbers to the right, text to the left
            if isinstance(cell.value, (int, float)):
                cell.alignment = right_alignment
            else:
                cell.alignment = left_alignment

    # Auto-adjust column widths with some padding
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)  # Add more padding
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Freeze the header row
    worksheet.freeze_panes = f"A{data_start_row + 1}"

# Helper function to calculate order price and profit
def calculate_order_price_and_profit(product: str, quantity: int) -> tuple:
    prices = {
        'Ortacha gulqand': 40000,  # Средний Гулканд
        'Katta gulqand': 50000     # Большой Гулканд
    }
    costs = {
        'Ortacha gulqand': 20000,  # Себестоимость среднего
        'Katta gulqand': 25000     # Себестоимость большого
    }
    price = prices.get(product, 0) * quantity
    cost = costs.get(product, 0) * quantity
    profit = price - cost
    return price, profit

# Helper function to generate Excel file
async def generate_excel_file():
    try:
        response = requests.get(BACKEND_URL)
        response.raise_for_status()
        orders = response.json()
        
        # Create DataFrame
        df = pd.DataFrame(orders)
        
        # Convert timestamp to readable format
        df['created_at'] = pd.to_datetime(df['created_at']).dt.strftime('%d.%m.%Y %H:%M:%S')
        
        # Calculate price and profit for each order
        df[['price', 'profit']] = df.apply(
            lambda x: pd.Series(calculate_order_price_and_profit(x['product'], x['quantity'])), 
            axis=1
        )
        
        # Reorder columns
        df = df[['id', 'name', 'phone', 'product', 'quantity', 'price', 'profit', 'status', 'created_at']]
        
        # Rename columns
        df.columns = ['ID', 'Имя', 'Телефон', 'Товар', 'Количество', 'Сумма', 'Прибыль', 'Статус', 'Дата создания']
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: All Orders
            df.to_excel(writer, index=False, sheet_name='Заказы')
            worksheet = writer.sheets['Заказы']
            apply_styles(worksheet, "Список всех заказов")

            # Sheet 2: Statistics
            # Calculate statistics
            total_orders = len(orders)
            approved_orders = len(df[df["Статус"] == "approved"])
            rejected_orders = len(df[df["Статус"] == "rejected"])
            pending_orders = len(df[df["Статус"] == "pending"])
            total_quantity = df["Количество"].sum()
            total_revenue = df[df["Статус"] == "approved"]["Сумма"].sum()
            total_profit = df[df["Статус"] == "approved"]["Прибыль"].sum()
            
            # Create statistics DataFrame
            stats_data = {
                'Показатель': [
                    'Всего заказов',
                    'Одобрено',
                    'Отклонено',
                    'Ожидает',
                    'Всего товаров',
                    'Общая выручка',
                    'Общая прибыль'
                ],
                'Значение': [
                    total_orders,
                    approved_orders,
                    rejected_orders,
                    pending_orders,
                    total_quantity,
                    f"{total_revenue:,.0f} сум",
                    f"{total_profit:,.0f} сум"
                ]
            }
            
            # Add product prices
            prices_data = {
                'Товар': ['Большой Гулканд', 'Средний Гулканд'],
                'Цена': ['50,000 сум', '40,000 сум'],
                'Себестоимость': ['25,000 сум', '20,000 сум'],
                'Маржа': ['25,000 сум', '20,000 сум']
            }
            
            # Add popular products
            product_stats = df.groupby('Товар').agg({
                'Количество': 'sum',
                'Сумма': 'sum',
                'Прибыль': 'sum'
            }).reset_index()
            product_stats = product_stats.sort_values('Количество', ascending=False)
            
            # Create statistics sheet
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, index=False, sheet_name='Статистика', startrow=0)
            
            # Add prices information
            prices_df = pd.DataFrame(prices_data)
            prices_df.to_excel(writer, index=False, sheet_name='Статистика', startrow=len(stats_data) + 3)
            
            # Add product statistics
            product_stats.to_excel(writer, index=False, sheet_name='Статистика', startrow=len(stats_data) + len(prices_data) + 6)
            
            worksheet = writer.sheets['Статистика']
            apply_styles(worksheet, "Статистика заказов")

            # Add pie chart for status distribution
            pie = PieChart()
            pie.title = "Распределение статусов заказов"
            pie.style = 10
            pie.height = 10
            pie.width = 15
            
            data_labels = DataLabelList()
            data_labels.showVal = True
            data_labels.showPercent = True
            pie.dLbls = data_labels
            
            # Create data for pie chart
            status_data = pd.DataFrame({
                'Статус': ['Одобрено', 'Отклонено', 'Ожидает'],
                'Количество': [approved_orders, rejected_orders, pending_orders]
            })
            status_data.to_excel(writer, index=False, sheet_name='Статистика', 
                               startrow=len(stats_data) + len(prices_data) + len(product_stats) + 9)
            
            data = Reference(worksheet, 
                           min_col=2, 
                           min_row=len(stats_data) + len(prices_data) + len(product_stats) + 10,
                           max_row=len(stats_data) + len(prices_data) + len(product_stats) + 12,
                           max_col=2)
            categories = Reference(worksheet,
                                min_col=1,
                                min_row=len(stats_data) + len(prices_data) + len(product_stats) + 10,
                                max_row=len(stats_data) + len(prices_data) + len(product_stats) + 12)
            
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(categories)
            worksheet.add_chart(pie, "E2")

            # Add pie chart for product distribution
            product_pie = PieChart()
            product_pie.title = "Распределение продаж по товарам"
            product_pie.style = 10
            product_pie.height = 10
            product_pie.width = 15
            
            product_data_labels = DataLabelList()
            product_data_labels.showVal = True
            product_data_labels.showPercent = True
            product_pie.dLbls = product_data_labels
            
            # Create data for product pie chart
            product_distribution = df[df['Статус'] == 'approved'].groupby('Товар')['Количество'].sum().reset_index()
            product_distribution.to_excel(writer, index=False, sheet_name='Статистика', 
                                        startrow=len(stats_data) + len(prices_data) + len(product_stats) + 15)
            
            product_data = Reference(worksheet,
                                  min_col=2,
                                  min_row=len(stats_data) + len(prices_data) + len(product_stats) + 16,
                                  max_row=len(stats_data) + len(prices_data) + len(product_stats) + 17,
                                  max_col=2)
            product_categories = Reference(worksheet,
                                        min_col=1,
                                        min_row=len(stats_data) + len(prices_data) + len(product_stats) + 16,
                                        max_row=len(stats_data) + len(prices_data) + len(product_stats) + 17)
            
            product_pie.add_data(product_data, titles_from_data=True)
            product_pie.set_categories(product_categories)
            worksheet.add_chart(product_pie, "E20")

            # Add bar chart for popular products
            chart = BarChart()
            chart.title = "Популярные товары"
            chart.y_axis.title = "Количество"
            chart.x_axis.title = "Товар"
            
            data = Reference(worksheet,
                           min_col=2,
                           min_row=len(stats_data) + len(prices_data) + 7,
                           max_row=len(stats_data) + len(prices_data) + len(product_stats) + 6,
                           max_col=2)
            categories = Reference(worksheet,
                                min_col=1,
                                min_row=len(stats_data) + len(prices_data) + 7,
                                max_row=len(stats_data) + len(prices_data) + len(product_stats) + 6)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            worksheet.add_chart(chart, "E38")
        
        output.seek(0)
        return output
    except Exception as e:
        print(f"Error generating Excel file: {e}")
        return None

# Handle "/start" command
@router.message(Command("start"))
async def handle_start(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Статистика", callback_data="view_stats")],
        [InlineKeyboardButton(text="📜 Одобренные", callback_data="view_approved_1")],
        [InlineKeyboardButton(text="🚫 Отклонённые", callback_data="view_rejected_1")],
        [InlineKeyboardButton(text="⏳ Ожидающие", callback_data="view_pending_1")],
        [InlineKeyboardButton(text="🔍 Поиск по ID", callback_data="search_by_id")],
        [InlineKeyboardButton(text="📱 Частые клиенты", callback_data="view_customers")],
        [InlineKeyboardButton(text="📥 Скачать статистику", callback_data="download_stats")],
        [InlineKeyboardButton(text="📅 Заказы за период", callback_data="select_period")],
        [InlineKeyboardButton(text="📈 Топ товары", callback_data="top_products")],
        [InlineKeyboardButton(text="💰 Финансовая сводка", callback_data="financial_summary")]
    ])

    await message.answer(
        "👋 Добро пожаловать в админ-панель!\n"
        "Выберите действие:",
        reply_markup=keyboard
    )

# Handle "/help" command
@router.message(Command("help"))
async def handle_help(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return

    help_text = "📝 Список доступных команд:\n\n"
    for cmd, desc in COMMANDS.items():
        help_text += f"/{cmd} - {desc}\n"

    await message.answer(help_text)

# Handle "/stats" command
@router.message(Command("stats"))
async def handle_stats_command(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return

    stats = await get_statistics()
    if not stats:
        await message.answer("❌ Ошибка при получении статистики")
        return

    message_text = (
        "📊 Статистика заказов:\n\n"
        f"📦 Всего заказов: {stats['total']}\n"
        f"✅ Одобрено: {stats['approved']}\n"
        f"❌ Отклонено: {stats['rejected']}\n"
        f"⏳ Ожидает: {stats['pending']}\n"
        f"📦 Всего товаров: {stats['total_quantity']}\n\n"
        "📈 Популярные товары:\n"
    )
    
    for product, quantity in sorted(stats['products'].items(), key=lambda x: x[1], reverse=True)[:5]:
        message_text += f"• {product}: {quantity} шт.\n"

    await message.answer(message_text)

# Handle "/orders" command
@router.message(Command("orders"))
async def handle_orders_command(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📜 Одобренные", callback_data="view_approved_1")],
        [InlineKeyboardButton(text="🚫 Отклонённые", callback_data="view_rejected_1")],
        [InlineKeyboardButton(text="⏳ Ожидающие", callback_data="view_pending_1")]
    ])

    await message.answer(
        "📋 Выберите тип заказов для просмотра:",
        reply_markup=keyboard
    )

# Handle "/pending" command
@router.message(Command("pending"))
async def handle_pending_command(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return
    
    try:
        response = requests.get(BACKEND_URL)
        response.raise_for_status()
        orders = response.json()
        
        # Filter pending orders
        pending_orders = [order for order in orders if order['status'] == 'pending']
        
        if not pending_orders:
            await message.answer(
                "❌ Нет ожидающих заказов",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
                ])
            )
            return
        
        # Show first page of pending orders
        page = 1
        orders_per_page = 5
        start_idx = (page - 1) * orders_per_page
        end_idx = start_idx + orders_per_page
        current_orders = pending_orders[start_idx:end_idx]
        
        message_text = f"📋 Ожидающие заказы (страница {page}):\n\n"
        
        for order in current_orders:
            price, profit = calculate_order_price_and_profit(order['product'], order['quantity'])
            message_text += (
                f"🆔 {order['id']}\n"
                f"👤 {order['name']}\n"
                f"📱 {order['phone']}\n"
                f"📦 {order['product']} x{order['quantity']}\n"
                f"💰 {price:,.0f} сум\n"
                f"📅 {datetime.fromisoformat(order['created_at'].replace('Z', '+00:00')).strftime('%d.%m.%Y %H:%M')}\n\n"
            )
        
        # Add navigation buttons
        keyboard = []
        if len(pending_orders) > orders_per_page:
            keyboard.append([
                InlineKeyboardButton(text="⬅️", callback_data=f"view_pending_{page-1}"),
                InlineKeyboardButton(text=f"{page}/{(len(pending_orders)-1)//orders_per_page + 1}", callback_data="page"),
                InlineKeyboardButton(text="➡️", callback_data=f"view_pending_{page+1}")
            ])
        
        keyboard.append([InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")])
        
        await message.answer(
            text=message_text,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard)
        )
        
    except Exception as e:
        print(f"Error handling pending orders: {e}")
        await message.answer(
            "❌ Ошибка при получении данных",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
            ])
        )

# Handle "/approved" command
@router.message(Command("approved"))
async def handle_approved_command(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return
    
    try:
        response = requests.get(BACKEND_URL)
        response.raise_for_status()
        orders = response.json()
        
        # Filter approved orders
        approved_orders = [order for order in orders if order['status'] == 'approved']
        
        if not approved_orders:
            await message.answer(
                "❌ Нет одобренных заказов",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
                ])
            )
            return
        
        # Show first page of approved orders
        page = 1
        orders_per_page = 5
        start_idx = (page - 1) * orders_per_page
        end_idx = start_idx + orders_per_page
        current_orders = approved_orders[start_idx:end_idx]
        
        message_text = f"📋 Одобренные заказы (страница {page}):\n\n"
        
        for order in current_orders:
            price, profit = calculate_order_price_and_profit(order['product'], order['quantity'])
            message_text += (
                f"🆔 {order['id']}\n"
                f"👤 {order['name']}\n"
                f"📱 {order['phone']}\n"
                f"📦 {order['product']} x{order['quantity']}\n"
                f"💰 {price:,.0f} сум\n"
                f"📅 {datetime.fromisoformat(order['created_at'].replace('Z', '+00:00')).strftime('%d.%m.%Y %H:%M')}\n\n"
            )
        
        # Add navigation buttons
        keyboard = []
        if len(approved_orders) > orders_per_page:
            keyboard.append([
                InlineKeyboardButton(text="⬅️", callback_data=f"view_approved_{page-1}"),
                InlineKeyboardButton(text=f"{page}/{(len(approved_orders)-1)//orders_per_page + 1}", callback_data="page"),
                InlineKeyboardButton(text="➡️", callback_data=f"view_approved_{page+1}")
            ])
        
        keyboard.append([InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")])
        
        await message.answer(
            text=message_text,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard)
        )
        
    except Exception as e:
        print(f"Error handling approved orders: {e}")
        await message.answer(
            "❌ Ошибка при получении данных",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
            ])
        )

# Handle "/rejected" command
@router.message(Command("rejected"))
async def handle_rejected_command(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return
    
    try:
        response = requests.get(BACKEND_URL)
        response.raise_for_status()
        orders = response.json()
        
        # Filter rejected orders
        rejected_orders = [order for order in orders if order['status'] == 'rejected']
        
        if not rejected_orders:
            await message.answer(
                "❌ Нет отклоненных заказов",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
                ])
            )
            return
        
        # Show first page of rejected orders
        page = 1
        orders_per_page = 5
        start_idx = (page - 1) * orders_per_page
        end_idx = start_idx + orders_per_page
        current_orders = rejected_orders[start_idx:end_idx]
        
        message_text = f"📋 Отклоненные заказы (страница {page}):\n\n"
        
        for order in current_orders:
            price, profit = calculate_order_price_and_profit(order['product'], order['quantity'])
            message_text += (
                f"🆔 {order['id']}\n"
                f"👤 {order['name']}\n"
                f"📱 {order['phone']}\n"
                f"📦 {order['product']} x{order['quantity']}\n"
                f"💰 {price:,.0f} сум\n"
                f"📅 {datetime.fromisoformat(order['created_at'].replace('Z', '+00:00')).strftime('%d.%m.%Y %H:%M')}\n\n"
            )
        
        # Add navigation buttons
        keyboard = []
        if len(rejected_orders) > orders_per_page:
            keyboard.append([
                InlineKeyboardButton(text="⬅️", callback_data=f"view_rejected_{page-1}"),
                InlineKeyboardButton(text=f"{page}/{(len(rejected_orders)-1)//orders_per_page + 1}", callback_data="page"),
                InlineKeyboardButton(text="➡️", callback_data=f"view_rejected_{page+1}")
            ])
        
        keyboard.append([InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")])
        
        await message.answer(
            text=message_text,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard)
        )
        
    except Exception as e:
        print(f"Error handling rejected orders: {e}")
        await message.answer(
            "❌ Ошибка при получении данных",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
            ])
        )

# Handle "/customers" command
@router.message(Command("customers"))
async def handle_customers_command(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return

    stats = await get_statistics()
    if not stats:
        await message.answer("❌ Ошибка при получении данных")
        return

    # Get top 10 customers
    top_customers = sorted(stats['customers'].items(), key=lambda x: x[1], reverse=True)[:10]
    
    message_text = "📱 Топ 10 частых клиентов:\n\n"
    for i, (phone, orders) in enumerate(top_customers, 1):
        message_text += f"{i}. {phone}: {orders} заказов\n"

    await message.answer(message_text)

# Handle "/finance" command
@router.message(Command("finance"))
async def handle_finance_command(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return
    
    try:
        response = requests.get(BACKEND_URL)
        response.raise_for_status()
        orders = response.json()
        
        # Calculate financial metrics
        total_revenue = 0
        total_profit = 0
        daily_revenue = defaultdict(int)
        daily_profit = defaultdict(int)
        product_stats = defaultdict(lambda: {'quantity': 0, 'revenue': 0, 'profit': 0})
        
        for order in orders:
            if order['status'] == 'approved':
                price, profit = calculate_order_price_and_profit(order['product'], order['quantity'])
                total_revenue += price
                total_profit += profit
                
                # Group by date
                date = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00')).strftime('%Y-%m-%d')
                daily_revenue[date] += price
                daily_profit[date] += profit
                
                # Group by product
                product_stats[order['product']]['quantity'] += order['quantity']
                product_stats[order['product']]['revenue'] += price
                product_stats[order['product']]['profit'] += profit
        
        # Calculate average daily revenue and profit
        avg_daily_revenue = total_revenue / len(daily_revenue) if daily_revenue else 0
        avg_daily_profit = total_profit / len(daily_profit) if daily_profit else 0
        
        message_text = (
            "💰 Финансовая сводка:\n\n"
            f"📈 Общая выручка: {total_revenue:,.0f} сум\n"
            f"💵 Общая прибыль: {total_profit:,.0f} сум\n"
            f"📊 Средняя дневная выручка: {avg_daily_revenue:,.0f} сум\n"
            f"📊 Средняя дневная прибыль: {avg_daily_profit:,.0f} сум\n\n"
            "📈 Статистика по товарам:\n"
        )
        
        # Add product statistics
        for product, stats in product_stats.items():
            message_text += (
                f"📦 {product}:\n"
                f"   • Количество: {stats['quantity']} шт.\n"
                f"   • Выручка: {stats['revenue']:,.0f} сум\n"
                f"   • Прибыль: {stats['profit']:,.0f} сум\n\n"
            )
        
        # Show last 7 days
        message_text += "📅 Последние 7 дней:\n"
        last_7_days = sorted(daily_revenue.items(), reverse=True)[:7]
        for date, revenue in last_7_days:
            profit = daily_profit[date]
            message_text += (
                f"📅 {date}: "
                f"💰 {revenue:,.0f} сум | "
                f"💵 {profit:,.0f} сум\n"
            )
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📥 Скачать детали", callback_data="download_financial")],
            [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
        ])
        
        await message.answer(
            text=message_text,
            reply_markup=keyboard
        )
        
    except Exception as e:
        print(f"Error handling financial summary: {e}")
        await message.answer(
            text="❌ Ошибка при получении данных.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
            ])
        )

# Handle "/products" command
@router.message(Command("products"))
async def handle_products_command(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return
    
    try:
        response = requests.get(BACKEND_URL)
        response.raise_for_status()
        orders = response.json()
        
        # Calculate product statistics
        product_stats = defaultdict(lambda: {'quantity': 0, 'revenue': 0, 'profit': 0})
        
        for order in orders:
            if order['status'] == 'approved':
                price, profit = calculate_order_price_and_profit(order['product'], order['quantity'])
                product_stats[order['product']]['quantity'] += order['quantity']
                product_stats[order['product']]['revenue'] += price
                product_stats[order['product']]['profit'] += profit
        
        # Sort by quantity
        sorted_products = sorted(product_stats.items(), key=lambda x: x[1]['quantity'], reverse=True)
        
        message_text = "📈 Топ товаров:\n\n"
        
        for product, stats in sorted_products:
            message_text += (
                f"📦 {product}:\n"
                f"   • Количество: {stats['quantity']} шт.\n"
                f"   • Выручка: {stats['revenue']:,.0f} сум\n"
                f"   • Прибыль: {stats['profit']:,.0f} сум\n"
                f"   • Средняя цена: {stats['revenue']/stats['quantity']:,.0f} сум\n\n"
            )
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📥 Скачать детали", callback_data="download_products")],
            [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
        ])
        
        await message.answer(
            text=message_text,
            reply_markup=keyboard
        )
        
    except Exception as e:
        print(f"Error handling top products: {e}")
        await message.answer(
            text="❌ Ошибка при получении данных.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
            ])
        )

# Handle "/download" command
@router.message(Command("download"))
async def handle_download_command(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        await message.answer("Вы не администратор!")
        return
    
    try:
        await message.answer("⏳ Генерация файла...")
        
        excel_file = await generate_excel_file()
        if not excel_file:
            await message.answer("❌ Ошибка при генерации файла статистики.")
            return

        # Send the Excel file
        file = BufferedInputFile(
            excel_file.getvalue(),
            filename=f"statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        await bot.send_document(
            chat_id=message.chat.id,
            document=file,
            caption="📊 Статистика заказов"
        )

        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
        ])
        await message.answer(
            text="✅ Файл статистики успешно сгенерирован",
            reply_markup=keyboard
        )
        
    except Exception as e:
        print(f"Error generating Excel file: {e}")
        await message.answer(
            text="❌ Ошибка при генерации файла.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
            ])
        )

# Handle search by ID
@router.callback_query(lambda c: c.data == "search_by_id")
async def handle_search_prompt(callback_query: CallbackQuery, bot: Bot):
    await bot.edit_message_text(
        chat_id=callback_query.message.chat.id,
        message_id=callback_query.message.message_id,
        text="🔍 Введите ID заказа для поиска:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
        ])
    )
    await bot.answer_callback_query(callback_query.id)

# Handle search by ID message
@router.message(F.text.regexp(r'^\d+$'))
async def handle_search_by_id(message: Message, bot: Bot):
    if message.chat.id not in ADMIN_IDS:
        return

    order_id = message.text
    try:
        response = requests.get(f"{BACKEND_URL}{order_id}/")
        response.raise_for_status()
        order = response.json()

        caption = (
            f"🔍 Результаты поиска по ID: {order_id}\n\n"
            f"🆔 ID: {order_id}\n"
            f"👤 Имя: {order['name']}\n"
            f"📅 Время: {format_timestamp(order['created_at'])}\n"
            f"📱 Телефон: {order['phone']}\n"
            f"📦 Товар: {order['product']}\n"
            f"🔢 Количество: {order['quantity']}\n"
            f"📝 Статус: {order['status']}\n"
        )

        keyboard_buttons = []
        if order['status'] == 'pending':
            keyboard_buttons.append([
                InlineKeyboardButton(text=f"✅ Одобрить {order_id}", callback_data=f"approve_{order_id}"),
                InlineKeyboardButton(text=f"❌ Отклонить {order_id}", callback_data=f"reject_{order_id}")
            ])

        base_url = BACKEND_URL.replace('/api/orders/', '')
        receipt_url = order['receipt'] if order['receipt'].startswith('http') else f"{base_url}{order['receipt']}"

        try:
            receipt_response = requests.get(receipt_url)
            receipt_response.raise_for_status()

            receipt_file = BufferedInputFile(receipt_response.content, filename=f"receipt_{order_id}.jpg")
            await bot.send_photo(
                chat_id=message.chat.id,
                photo=receipt_file,
                caption=caption,
                reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard_buttons) if keyboard_buttons else None
            )
        except requests.RequestException as e:
            print(f"Ошибка при загрузке чека: {e} для Order ID: {order_id}")
            await bot.send_message(
                chat_id=message.chat.id,
                text=f"{caption}\n❌ Ошибка: Не удалось загрузить чек.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard_buttons) if keyboard_buttons else None
            )
        except Exception as e:
            print(f"Ошибка при отправке фото: {e} для Order ID: {order_id}")
            await bot.send_message(
                chat_id=message.chat.id,
                text=f"{caption}\n❌ Ошибка: Не удалось отправить чек.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard_buttons) if keyboard_buttons else None
            )

        # Add back button
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
        ])
        await bot.send_message(
            chat_id=message.chat.id,
            text="🔍 Поиск завершен",
            reply_markup=keyboard
        )

    except requests.RequestException as e:
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
        ])
        await bot.send_message(
            chat_id=message.chat.id,
            text=f"❌ Заказ с ID {order_id} не найден.",
            reply_markup=keyboard
        )

# Handle "Approve" / "Reject" button presses
@router.callback_query(lambda c: c.data.startswith(('approve_', 'reject_')))
async def handle_approval(callback_query: CallbackQuery, bot: Bot):
    action, order_id = callback_query.data.split('_', 1)
    status = "approved" if action == "approve" else "rejected"

    try:
        response = requests.patch(
            f"{BACKEND_URL}{order_id}/",
            json={'status': status},
            headers={'Content-Type': 'application/json'}
        )
        response.raise_for_status()

        await bot.answer_callback_query(callback_query.id, f"✅ Статус изменён на: {status}")
        await bot.edit_message_reply_markup(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            reply_markup=None
        )
        if order_id in sent_order_ids:
            sent_order_ids.remove(order_id)
    except requests.RequestException as e:
        error_msg = f"Ошибка обновления статуса: {str(e)} - Response: {e.response.text if e.response else 'No response'}"
        print(error_msg)
        await bot.answer_callback_query(callback_query.id, "❌ Ошибка при обновлении статуса")

# Handlers for viewing orders by status with pagination
@router.callback_query(lambda c: c.data.startswith(('view_approved_', 'view_rejected_', 'view_pending_', 'back_to_main')))
async def handle_view_orders(callback_query: CallbackQuery, bot: Bot):
    if callback_query.data == "back_to_main":
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📊 Статистика", callback_data="view_stats")],
            [InlineKeyboardButton(text="📜 Одобренные", callback_data="view_approved_1")],
            [InlineKeyboardButton(text="🚫 Отклонённые", callback_data="view_rejected_1")],
            [InlineKeyboardButton(text="⏳ Ожидающие", callback_data="view_pending_1")],
            [InlineKeyboardButton(text="🔍 Поиск по ID", callback_data="search_by_id")],
            [InlineKeyboardButton(text="📱 Частые клиенты", callback_data="view_customers")],
            [InlineKeyboardButton(text="📥 Скачать статистику", callback_data="download_stats")],
            [InlineKeyboardButton(text="📅 Заказы за период", callback_data="select_period")],
            [InlineKeyboardButton(text="📈 Топ товары", callback_data="top_products")],
            [InlineKeyboardButton(text="💰 Финансовая сводка", callback_data="financial_summary")]
        ])
        await bot.edit_message_text(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            text="👋 Добро пожаловать в админ-панель!\nВыберите действие:",
            reply_markup=keyboard
        )
        return

    # Split the callback data correctly
    parts = callback_query.data.split('_')
    status = f"{parts[0]}_{parts[1]}"  # e.g., "view_approved"
    page = int(parts[2])  # e.g., "1"

    status_map = {
        'view_approved': 'approved',
        'view_rejected': 'rejected',
        'view_pending': 'pending'
    }
    current_status = status_map[status]

    try:
        response = requests.get(f"{BACKEND_URL}?status={current_status}")
        response.raise_for_status()
        orders = response.json()

        if not orders:
            await bot.edit_message_text(
                chat_id=callback_query.message.chat.id,
                message_id=callback_query.message.message_id,
                text=f"❌ Нет заказов со статусом '{current_status}'.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
                ])
            )
            return

        total_pages = (len(orders) + ORDERS_PER_PAGE - 1) // ORDERS_PER_PAGE
        start_idx = (page - 1) * ORDERS_PER_PAGE
        end_idx = min(start_idx + ORDERS_PER_PAGE, len(orders))
        current_orders = orders[start_idx:end_idx]

        base_url = BACKEND_URL.replace('/api/orders/', '')
        
        for order in current_orders:
            order_id = str(order['id'])
            caption = (
                f"📋 Заказ со статусом '{current_status}'\n\n"
                f"🆔 ID: {order_id}\n"
                f"👤 Имя: {order['name']}\n"
                f"📅 Время: {format_timestamp(order['created_at'])}\n"
                f"📱 Телефон: {order['phone']}\n"
                f"📦 Товар: {order['product']}\n"
                f"🔢 Количество: {order['quantity']}\n"
                f"📝 Статус: {order['status']}\n"
            )

            keyboard_buttons = []
            if current_status == 'pending':
                keyboard_buttons.append([
                    InlineKeyboardButton(text=f"✅ Одобрить {order_id}", callback_data=f"approve_{order_id}"),
                    InlineKeyboardButton(text=f"❌ Отклонить {order_id}", callback_data=f"reject_{order_id}")
                ])

            receipt_url = order['receipt'] if order['receipt'].startswith('http') else f"{base_url}{order['receipt']}"

            try:
                receipt_response = requests.get(receipt_url)
                receipt_response.raise_for_status()

                receipt_file = BufferedInputFile(receipt_response.content, filename=f"receipt_{order_id}.jpg")
                await bot.send_photo(
                    chat_id=callback_query.message.chat.id,
                    photo=receipt_file,
                    caption=caption,
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard_buttons) if keyboard_buttons else None
                )
            except requests.RequestException as e:
                print(f"Ошибка при загрузке чека: {e} для Order ID: {order_id}")
                await bot.send_message(
                    chat_id=callback_query.message.chat.id,
                    text=f"{caption}\n❌ Ошибка: Не удалось загрузить чек.",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard_buttons) if keyboard_buttons else None
                )
            except Exception as e:
                print(f"Ошибка при отправке фото: {e} для Order ID: {order_id}")
                await bot.send_message(
                    chat_id=callback_query.message.chat.id,
                    text=f"{caption}\n❌ Ошибка: Не удалось отправить чек.",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard_buttons) if keyboard_buttons else None
                )

        # Add pagination buttons
        pagination_buttons = []
        if page > 1:
            pagination_buttons.append(InlineKeyboardButton(text="⬅️ Назад", callback_data=f"{status}_{page-1}"))
        if page < total_pages:
            pagination_buttons.append(InlineKeyboardButton(text="➡️ Вперед", callback_data=f"{status}_{page+1}"))
        
        if pagination_buttons:
            keyboard_buttons = [pagination_buttons]
            keyboard_buttons.append([InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")])
            await bot.send_message(
                chat_id=callback_query.message.chat.id,
                text=f"📄 Страница {page} из {total_pages}",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard_buttons)
            )
        else:
            keyboard_buttons = [[InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]]
            await bot.send_message(
                chat_id=callback_query.message.chat.id,
                text="📄 Конец списка",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard_buttons)
            )

        await bot.answer_callback_query(callback_query.id)
    except requests.RequestException as e:
        print(f"Ошибка при получении заказов: {e}")
        await bot.edit_message_text(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            text=f"❌ Ошибка при загрузке заказов со статусом '{current_status}'.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
            ])
        )
        await bot.answer_callback_query(callback_query.id)

# Handle frequent customers view
@router.callback_query(lambda c: c.data == "view_customers")
async def handle_customers(callback_query: CallbackQuery, bot: Bot):
    stats = await get_statistics()
    if not stats:
        await bot.answer_callback_query(callback_query.id, "Ошибка при получении данных")
        return

    # Get top 10 customers
    top_customers = sorted(stats['customers'].items(), key=lambda x: x[1], reverse=True)[:10]
    
    message = "📱 Топ 10 частых клиентов:\n\n"
    for i, (phone, orders) in enumerate(top_customers, 1):
        message += f"{i}. {phone}: {orders} заказов\n"

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
    ])

    await bot.edit_message_text(
        chat_id=callback_query.message.chat.id,
        message_id=callback_query.message.message_id,
        text=message,
        reply_markup=keyboard
    )
    await bot.answer_callback_query(callback_query.id)

# Handle Excel download
@router.callback_query(lambda c: c.data == "download_stats")
async def handle_download_stats(callback_query: CallbackQuery, bot: Bot):
    await bot.answer_callback_query(callback_query.id, "⏳ Генерация файла...")
    
    excel_file = await generate_excel_file()
    if not excel_file:
        await bot.send_message(
            chat_id=callback_query.message.chat.id,
            text="❌ Ошибка при генерации файла статистики."
        )
        return

    # Send the Excel file
    file = BufferedInputFile(
        excel_file.getvalue(),
        filename=f"statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    
    await bot.send_document(
        chat_id=callback_query.message.chat.id,
        document=file,
        caption="📊 Статистика заказов"
    )

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
    ])
    await bot.send_message(
        chat_id=callback_query.message.chat.id,
        text="✅ Файл статистики успешно сгенерирован",
        reply_markup=keyboard
    )

# Handle statistics view
@router.callback_query(lambda c: c.data == "view_stats")
async def handle_statistics(callback_query: CallbackQuery, bot: Bot):
    stats = await get_statistics()
    if not stats:
        await bot.answer_callback_query(callback_query.id, "Ошибка при получении статистики")
        return

    message = (
        "📊 Статистика заказов:\n\n"
        f"📦 Всего заказов: {stats['total']}\n"
        f"✅ Одобрено: {stats['approved']}\n"
        f"❌ Отклонено: {stats['rejected']}\n"
        f"⏳ Ожидает: {stats['pending']}\n"
        f"📦 Всего товаров: {stats['total_quantity']}\n\n"
        "📈 Популярные товары:\n"
    )
    
    for product, quantity in sorted(stats['products'].items(), key=lambda x: x[1], reverse=True)[:5]:
        message += f"• {product}: {quantity} шт.\n"

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
    ])

    await bot.edit_message_text(
        chat_id=callback_query.message.chat.id,
        message_id=callback_query.message.message_id,
        text=message,
        reply_markup=keyboard
    )
    await bot.answer_callback_query(callback_query.id)

# Function to send order to Telegram admin
async def send_order_to_admin(bot: Bot, order):
    order_id = str(order['id'])
    if order_id in sent_order_ids:
        return

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Одобрить", callback_data=f"approve_{order_id}")],
        [InlineKeyboardButton(text="❌ Отклонить", callback_data=f"reject_{order_id}")]
    ])

    caption = (
        f"🛒 Новый заказ!\n\n"
        f"🆔 ID: {order_id}\n"
        f"👤 Имя: {order['name']}\n"
        f"📅 Время: {format_timestamp(order['created_at'])}\n"
        f"📱 Телефон: {order['phone']}\n"
        f"📦 Товар: {order['product']}\n"
        f"🔢 Количество: {order['quantity']}\n"
    )

    base_url = BACKEND_URL.replace('/api/orders/', '')
    receipt_url = order['receipt'] if order['receipt'].startswith('http') else f"{base_url}{order['receipt']}"

    try:
        receipt_response = requests.get(receipt_url)
        receipt_response.raise_for_status()

        receipt_file = BufferedInputFile(receipt_response.content, filename=f"receipt_{order_id}.jpg")
        await bot.send_photo(
            chat_id=ADMIN_CHAT_ID,
            photo=receipt_file,
            caption=caption,
            reply_markup=keyboard
        )
        sent_order_ids.add(order_id)
    except requests.RequestException as e:
        print(f"Ошибка при загрузке чека: {e} для Order ID: {order_id}")
        await bot.send_message(
            chat_id=ADMIN_CHAT_ID,
            text=f"{caption}\n❌ Ошибка: Не удалось загрузить чек.",
            reply_markup=keyboard
        )
        sent_order_ids.add(order_id)
    except Exception as e:
        print(f"Ошибка при отправке фото: {e} для Order ID: {order_id}")
        await bot.send_message(
            chat_id=ADMIN_CHAT_ID,
            text=f"{caption}\n❌ Ошибка: Не удалось отправить чек.",
            reply_markup=keyboard
        )
        sent_order_ids.add(order_id)

# Handle period selection
@router.callback_query(lambda c: c.data == "select_period")
async def handle_period_selection(callback_query: CallbackQuery, bot: Bot):
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📅 Сегодня", callback_data="period_today")],
        [InlineKeyboardButton(text="📅 Вчера", callback_data="period_yesterday")],
        [InlineKeyboardButton(text="📅 Неделя", callback_data="period_week")],
        [InlineKeyboardButton(text="📅 Месяц", callback_data="period_month")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
    ])
    
    await bot.edit_message_text(
        chat_id=callback_query.message.chat.id,
        message_id=callback_query.message.message_id,
        text="Выберите период для просмотра заказов:",
        reply_markup=keyboard
    )

# Handle period selection
@router.callback_query(lambda c: c.data.startswith("period_"))
async def handle_period_orders(callback_query: CallbackQuery, bot: Bot):
    period = callback_query.data.split("_")[1]
    now = datetime.now()
    
    if period == "today":
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    elif period == "yesterday":
        start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start_date = now - timedelta(days=7)
        end_date = now
    elif period == "month":
        start_date = now - timedelta(days=30)
        end_date = now
    
    try:
        response = requests.get(BACKEND_URL)
        response.raise_for_status()
        orders = response.json()
        
        # Filter orders by date
        filtered_orders = []
        for order in orders:
            # Convert order date to naive datetime for comparison
            order_date = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))
            order_date = order_date.replace(tzinfo=None)  # Make it naive
            
            if start_date <= order_date <= end_date:
                filtered_orders.append(order)
        
        if not filtered_orders:
            await bot.edit_message_text(
                chat_id=callback_query.message.chat.id,
                message_id=callback_query.message.message_id,
                text=f"❌ Нет заказов за выбранный период.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🔙 Назад", callback_data="select_period")]
                ])
            )
            return
        
        # Calculate statistics
        total_orders = len(filtered_orders)
        total_quantity = sum(order['quantity'] for order in filtered_orders)
        total_revenue = sum(calculate_order_price_and_profit(order['product'], order['quantity'])[0] 
                          for order in filtered_orders if order['status'] == 'approved')
        total_profit = sum(calculate_order_price_and_profit(order['product'], order['quantity'])[1] 
                         for order in filtered_orders if order['status'] == 'approved')
        
        # Group by product
        product_stats = defaultdict(lambda: {'quantity': 0, 'revenue': 0, 'profit': 0})
        for order in filtered_orders:
            if order['status'] == 'approved':
                price, profit = calculate_order_price_and_profit(order['product'], order['quantity'])
                product_stats[order['product']]['quantity'] += order['quantity']
                product_stats[order['product']]['revenue'] += price
                product_stats[order['product']]['profit'] += profit
        
        message = (
            f"📊 Статистика за {period}:\n\n"
            f"📦 Всего заказов: {total_orders}\n"
            f"📦 Всего товаров: {total_quantity}\n"
            f"💰 Выручка: {total_revenue:,.0f} сум\n"
            f"💵 Прибыль: {total_profit:,.0f} сум\n\n"
            "📈 Статистика по товарам:\n"
        )
        
        # Add product statistics
        for product, stats in product_stats.items():
            message += (
                f"📦 {product}:\n"
                f"   • Количество: {stats['quantity']} шт.\n"
                f"   • Выручка: {stats['revenue']:,.0f} сум\n"
                f"   • Прибыль: {stats['profit']:,.0f} сум\n\n"
            )
        
        # Show last 5 orders
        message += "📋 Последние заказы:\n"
        for order in filtered_orders[-5:]:
            price, profit = calculate_order_price_and_profit(order['product'], order['quantity'])
            message += (
                f"🆔 {order['id']} | "
                f"📦 {order['product']} x{order['quantity']} | "
                f"💰 {price:,.0f} сум | "
                f"📝 {order['status']}\n"
            )
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📥 Скачать детали", callback_data=f"download_period_{period}")],
            [InlineKeyboardButton(text="🔙 Назад", callback_data="select_period")]
        ])
        
        await bot.edit_message_text(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            text=message,
            reply_markup=keyboard
        )
        
    except Exception as e:
        print(f"Error handling period orders: {e}")
        await bot.edit_message_text(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            text="❌ Ошибка при получении данных.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Назад", callback_data="select_period")]
            ])
        )

# Handle financial summary
@router.callback_query(lambda c: c.data == "financial_summary")
async def handle_financial_summary(callback_query: CallbackQuery, bot: Bot):
    try:
        response = requests.get(BACKEND_URL)
        response.raise_for_status()
        orders = response.json()
        
        # Calculate financial metrics
        total_revenue = 0
        total_profit = 0
        daily_revenue = defaultdict(int)
        daily_profit = defaultdict(int)
        product_stats = defaultdict(lambda: {'quantity': 0, 'revenue': 0, 'profit': 0})
        
        for order in orders:
            if order['status'] == 'approved':
                price, profit = calculate_order_price_and_profit(order['product'], order['quantity'])
                total_revenue += price
                total_profit += profit
                
                # Group by date
                date = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00')).strftime('%Y-%m-%d')
                daily_revenue[date] += price
                daily_profit[date] += profit
                
                # Group by product
                product_stats[order['product']]['quantity'] += order['quantity']
                product_stats[order['product']]['revenue'] += price
                product_stats[order['product']]['profit'] += profit
        
        # Calculate average daily revenue and profit
        avg_daily_revenue = total_revenue / len(daily_revenue) if daily_revenue else 0
        avg_daily_profit = total_profit / len(daily_profit) if daily_profit else 0
        
        message = (
            "💰 Финансовая сводка:\n\n"
            f"📈 Общая выручка: {total_revenue:,.0f} сум\n"
            f"💵 Общая прибыль: {total_profit:,.0f} сум\n"
            f"📊 Средняя дневная выручка: {avg_daily_revenue:,.0f} сум\n"
            f"📊 Средняя дневная прибыль: {avg_daily_profit:,.0f} сум\n\n"
            "📈 Статистика по товарам:\n"
        )
        
        # Add product statistics
        for product, stats in product_stats.items():
            message += (
                f"📦 {product}:\n"
                f"   • Количество: {stats['quantity']} шт.\n"
                f"   • Выручка: {stats['revenue']:,.0f} сум\n"
                f"   • Прибыль: {stats['profit']:,.0f} сум\n\n"
            )
        
        # Show last 7 days
        message += "📅 Последние 7 дней:\n"
        last_7_days = sorted(daily_revenue.items(), reverse=True)[:7]
        for date, revenue in last_7_days:
            profit = daily_profit[date]
            message += (
                f"📅 {date}: "
                f"💰 {revenue:,.0f} сум | "
                f"💵 {profit:,.0f} сум\n"
            )
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📥 Скачать детали", callback_data="download_financial")],
            [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
        ])
        
        await bot.edit_message_text(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            text=message,
            reply_markup=keyboard
        )
        
    except Exception as e:
        print(f"Error handling financial summary: {e}")
        await bot.edit_message_text(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            text="❌ Ошибка при получении данных.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
            ])
        )

# Handle top products
@router.callback_query(lambda c: c.data == "top_products")
async def handle_top_products(callback_query: CallbackQuery, bot: Bot):
    try:
        response = requests.get(BACKEND_URL)
        response.raise_for_status()
        orders = response.json()
        
        # Calculate product statistics
        product_stats = defaultdict(lambda: {'quantity': 0, 'revenue': 0, 'profit': 0})
        
        for order in orders:
            if order['status'] == 'approved':
                price, profit = calculate_order_price_and_profit(order['product'], order['quantity'])
                product_stats[order['product']]['quantity'] += order['quantity']
                product_stats[order['product']]['revenue'] += price
                product_stats[order['product']]['profit'] += profit
        
        # Sort by quantity
        sorted_products = sorted(product_stats.items(), key=lambda x: x[1]['quantity'], reverse=True)
        
        message = "📈 Топ товаров:\n\n"
        
        for product, stats in sorted_products:
            message += (
                f"📦 {product}:\n"
                f"   • Количество: {stats['quantity']} шт.\n"
                f"   • Выручка: {stats['revenue']:,.0f} сум\n"
                f"   • Прибыль: {stats['profit']:,.0f} сум\n"
                f"   • Средняя цена: {stats['revenue']/stats['quantity']:,.0f} сум\n\n"
            )
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📥 Скачать детали", callback_data="download_products")],
            [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
        ])
        
        await bot.edit_message_text(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            text=message,
            reply_markup=keyboard
        )
        
    except Exception as e:
        print(f"Error handling top products: {e}")
        await bot.edit_message_text(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            text="❌ Ошибка при получении данных.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
            ])
        )